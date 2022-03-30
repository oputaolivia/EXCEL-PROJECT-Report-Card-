import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('ReportCard.xlsx')
sheet = wb['Sheet1']

for row in range(4, sheet.max_row):
    cell = sheet.cell(row, 2)
    cell2 = sheet.cell(row, 3)
    summation = cell.value + cell2.value
    summation_cell = sheet.cell(row, 4)
    summation_cell.value = summation

    if summation >= 80:
        grade = 'A'
        print(grade)
        grade_cell = sheet.cell(row, 5)
        grade_cell.value = grade
    elif summation >= 60:
        grade = 'B'
        print(grade)
        grade_cell = sheet.cell(row, 5)
        grade_cell.value = grade
    elif summation >= 40:
        grade ='C'
        print(grade)
        grade_cell = sheet.cell(row, 5)
        grade_cell.value = grade
    elif summation >= 20:
        grade = 'D'
        print(grade)
        grade_cell = sheet.cell(row, 5)
        grade_cell.value = grade
    else:
        grade = 'F'
        print(grade)
        grade_cell = sheet.cell(row, 5)
        grade_cell.value = grade

    if summation >= 40:
        grade_pf = 'PASS'
        print(grade_pf)
        grade_pf_cell = sheet.cell(row, 6)
        grade_pf_cell.value = grade_pf
    else:
        grade_pf = 'FAIL'
        print(grade_pf)
        grade_pf_cell = sheet.cell(row, 6)
        grade_pf_cell.value = grade_pf


values = Reference(sheet,
                   min_row=4,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
labels = Reference(sheet,
                   min_row=4,
                   max_row=sheet.max_row,
                   min_col=1)
chart.add_data(values)
chart.set_categories(labels)
chart.style = 47
sheet.add_chart(chart, 'a18')


wb.save('ReportCard2.xlsx')
